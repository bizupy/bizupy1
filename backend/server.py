from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Request, status, Response, Cookie
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import base64
import io
from PIL import Image
import PyPDF2
import json
import jwt
from passlib.context import CryptContext
import random
import razorpay
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
import asyncio
import httpx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT & Password hashing
JWT_SECRET = os.environ.get('JWT_SECRET', 'change_me_in_production')
JWT_ALGORITHM = 'HS256'
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Razorpay client
razorpay_key_id = os.environ.get('RAZORPAY_KEY_ID', '')
razorpay_key_secret = os.environ.get('RAZORPAY_KEY_SECRET', '')
razorpay_client = razorpay.Client(auth=(razorpay_key_id, razorpay_key_secret)) if razorpay_key_id else None

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / 'uploads'
UPLOADS_DIR.mkdir(exist_ok=True)

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserBase(BaseModel):
    email: EmailStr
    name: str
    picture: Optional[str] = None
    language_preference: str = "en"
    business_name: Optional[str] = None
    business_gstin: Optional[str] = None
    business_address: Optional[str] = None
    business_phone: Optional[str] = None
    business_logo: Optional[str] = None

class UserResponse(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    language_preference: str
    business_name: Optional[str] = None
    business_gstin: Optional[str] = None
    business_logo: Optional[str] = None
    subscription_plan: str = "free"
    created_at: str

class UserProfileUpdate(BaseModel):
    name: Optional[str] = None
    language_preference: Optional[str] = None
    business_name: Optional[str] = None
    business_gstin: Optional[str] = None
    business_address: Optional[str] = None
    business_phone: Optional[str] = None
    business_logo: Optional[str] = None

class GoogleSessionRequest(BaseModel):
    session_id: str

class BillExtractedData(BaseModel):
    seller_gstin: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_gstin: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    products: List[Dict[str, Any]] = []
    subtotal: Optional[float] = None
    cgst: Optional[float] = None
    sgst: Optional[float] = None
    igst: Optional[float] = None
    total_gst: Optional[float] = None
    total_amount: Optional[float] = None
    confidence_score: float = 0.0

class BillResponse(BaseModel):
    id: str
    user_id: str
    file_name: str
    file_type: str
    upload_date: str
    ocr_status: str
    extracted_data: Optional[BillExtractedData] = None

class CustomerBase(BaseModel):
    name: str
    gstin: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None

class CustomerResponse(CustomerBase):
    id: str
    user_id: str
    total_purchases: float = 0.0
    created_at: str

class ProductBase(BaseModel):
    name: str
    hsn_code: Optional[str] = None
    unit: str = "pcs"
    default_price: float = 0.0

class ProductResponse(ProductBase):
    id: str
    user_id: str
    created_at: str

class InvoiceItem(BaseModel):
    product_name: str
    hsn_code: Optional[str] = None
    quantity: float
    unit: str = "pcs"
    rate: float
    amount: float

class InvoiceCreate(BaseModel):
    customer_id: Optional[str] = None
    customer_name: str
    customer_gstin: Optional[str] = None
    customer_address: Optional[str] = None
    items: List[InvoiceItem]
    notes: Optional[str] = None

class InvoiceResponse(BaseModel):
    id: str
    user_id: str
    invoice_number: str
    invoice_date: str
    customer_name: str
    customer_gstin: Optional[str] = None
    items: List[InvoiceItem]
    subtotal: float
    cgst: float
    sgst: float
    igst: float
    total_gst: float
    total_amount: float
    created_at: str

class DashboardStats(BaseModel):
    total_bills: int
    total_customers: int
    total_sales: float
    total_gst: float
    monthly_sales: float
    monthly_gst: float
    recent_bills: List[BillResponse]

class SubscriptionOrder(BaseModel):
    plan: str
    billing_cycle: str = "monthly"

# ==================== AUTHENTICATION ====================

async def get_current_user_from_token(token: str) -> dict:
    """Get user from session token"""
    try:
        # Find session in database
        session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
        if not session:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        # Check expiry
        expires_at = session["expires_at"]
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if expires_at < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Session expired")
        
        # Get user
        user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        return user
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verifying session: {str(e)}")
        raise HTTPException(status_code=401, detail="Invalid session")

async def get_current_user(
    request: Request,
    session_token: Optional[str] = Cookie(None)
) -> dict:
    """Get current user from cookie or Authorization header"""
    # Try cookie first
    if session_token:
        return await get_current_user_from_token(session_token)
    
    # Try Authorization header
    auth_header = request.headers.get('Authorization')
    if auth_header and auth_header.startswith('Bearer '):
        token = auth_header.split(' ')[1]
        return await get_current_user_from_token(token)
    
    raise HTTPException(status_code=401, detail="Not authenticated")

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/google-session")
async def process_google_session(session_req: GoogleSessionRequest, response: Response):
    """Process Google session_id and create user session"""
    try:
        logger.info(f"Processing Google session: {session_req.session_id[:20]}...")
        
        # Call Emergent Auth API to get user data
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_req.session_id},
                timeout=10.0
            )
            
            logger.info(f"Emergent Auth API response status: {auth_response.status_code}")
            
            if auth_response.status_code != 200:
                logger.error(f"Emergent Auth API error: {auth_response.text}")
                raise HTTPException(
                    status_code=400, 
                    detail=f"Invalid session ID or authentication failed: {auth_response.text}"
                )
            
            user_data = auth_response.json()
            logger.info(f"User data received for: {user_data.get('email', 'unknown')}")
        
        # Validate required fields
        if not user_data.get('email'):
            raise HTTPException(status_code=400, detail="No email in user data")
        if not user_data.get('session_token'):
            raise HTTPException(status_code=400, detail="No session token in user data")
        
        # Check if user exists
        existing_user = await db.users.find_one({"email": user_data["email"]}, {"_id": 0})
        
        if existing_user:
            user_id = existing_user["user_id"]
            logger.info(f"Existing user found: {user_id}")
            # Update user info
            await db.users.update_one(
                {"user_id": user_id},
                {"$set": {
                    "name": user_data.get("name", existing_user.get("name")),
                    "picture": user_data.get("picture"),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        else:
            # Create new user with custom user_id
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            logger.info(f"Creating new user: {user_id}")
            new_user = {
                "user_id": user_id,
                "email": user_data["email"],
                "name": user_data.get("name", user_data["email"].split('@')[0]),
                "picture": user_data.get("picture"),
                "language_preference": "en",
                "subscription_plan": "free",
                "bill_count": 0,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(new_user)
        
        # Create session
        session_token = user_data["session_token"]
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        # Delete old sessions for this user
        await db.user_sessions.delete_many({"user_id": user_id})
        
        session_doc = {
            "user_id": user_id,
            "session_token": session_token,
            "expires_at": expires_at.isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.user_sessions.insert_one(session_doc)
        logger.info(f"Session created for user: {user_id}")
        
        # Set httpOnly cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=False,  # Set to True in production with HTTPS
            samesite="lax",
            path="/",
            max_age=7 * 24 * 60 * 60  # 7 days
        )
        
        # Get full user data
        user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        
        logger.info(f"Authentication successful for: {user['email']}")
        
        return {
            "user": UserResponse(**user),
            "session_token": session_token,
            "message": "Authentication successful"
        }
        
    except httpx.RequestError as e:
        logger.error(f"Error calling Emergent Auth API: {str(e)}")
        raise HTTPException(
            status_code=503, 
            detail="Authentication service unavailable. Please try again later."
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing Google session: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500, 
            detail=f"Authentication error: {str(e)}"
        )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    """Get current user info"""
    return UserResponse(**user)

@api_router.post("/auth/logout")
async def logout(
    response: Response,
    user: dict = Depends(get_current_user),
    session_token: Optional[str] = Cookie(None)
):
    """Logout user"""
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.post("/auth/upload-logo")
async def upload_logo(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload business logo"""
    allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Only JPG, PNG, and WEBP are allowed.")
    
    try:
        file_content = await file.read()
        
        # Optimize and resize image
        img = Image.open(io.BytesIO(file_content))
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Resize to max 400x400 for logo
        max_size = 400
        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        
        # Save optimized logo
        logo_id = f"logo_{user['user_id']}"
        logo_path = UPLOADS_DIR / f"{logo_id}.jpg"
        img.save(logo_path, format='JPEG', quality=90, optimize=True)
        
        # Update user profile with logo path
        logo_url = f"/uploads/{logo_id}.jpg"
        await db.users.update_one(
            {"user_id": user['user_id']},
            {"$set": {"business_logo": logo_url}}
        )
        
        return {"message": "Logo uploaded successfully", "logo_url": logo_url}
        
    except Exception as e:
        logger.error(f"Error uploading logo: {str(e)}")
        raise HTTPException(status_code=500, detail="Error uploading logo")

@api_router.put("/auth/profile")
async def update_profile(profile: UserProfileUpdate, user: dict = Depends(get_current_user)):
    """Update user profile"""
    update_data = profile.model_dump(exclude_unset=True)
    await db.users.update_one({"user_id": user['user_id']}, {"$set": update_data})
    return {"message": "Profile updated successfully"}

# ==================== BILL PROCESSING ====================

async def extract_bill_data(image_base64: str) -> BillExtractedData:
    """Extract bill data using OpenAI GPT-5.2 vision"""
    try:
        emergent_key = os.environ.get('EMERGENT_LLM_KEY')
        if not emergent_key:
            raise Exception("EMERGENT_LLM_KEY not found")
        
        chat = LlmChat(
            api_key=emergent_key,
            session_id=str(uuid.uuid4()),
            system_message="""You are an expert Indian GST bill data extractor. Extract all relevant information from the bill image and return it as a JSON object. 
            Extract: seller_gstin, seller_name, buyer_gstin, buyer_name, invoice_number, invoice_date, products (array with name, hsn_code, quantity, rate, amount), subtotal, cgst, sgst, igst, total_gst, total_amount.
            If any field is not found, use null. Also provide a confidence_score (0-1) based on image quality and data clarity.
            Return ONLY valid JSON, no markdown or extra text."""
        ).with_model("openai", "gpt-5.2")
        
        image_content = ImageContent(image_base64=image_base64)
        user_message = UserMessage(
            text="Extract all GST bill information from this image and return as JSON.",
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        
        # Parse JSON response
        try:
            response_text = response.strip()
            if response_text.startswith('```'):
                response_text = response_text.split('\n', 1)[1].rsplit('\n', 1)[0].strip()
                if response_text.startswith('json'):
                    response_text = response_text[4:].strip()
            
            data = json.loads(response_text)
            return BillExtractedData(**data)
        except json.JSONDecodeError:
            logger.error(f"Failed to parse JSON response: {response}")
            return BillExtractedData(confidence_score=0.0)
            
    except Exception as e:
        logger.error(f"Error extracting bill data: {str(e)}")
        return BillExtractedData(confidence_score=0.0)

def convert_to_base64(file_content: bytes, file_type: str) -> str:
    """Convert image or PDF to base64"""
    try:
        if file_type.lower() == 'pdf':
            return base64.b64encode(file_content).decode('utf-8')
        else:
            img = Image.open(io.BytesIO(file_content))
            if img.mode != 'RGB':
                img = img.convert('RGB')
            max_size = 2048
            if max(img.size) > max_size:
                ratio = max_size / max(img.size)
                new_size = tuple(int(dim * ratio) for dim in img.size)
                img = img.resize(new_size, Image.Resampling.LANCZOS)
            buffered = io.BytesIO()
            img.save(buffered, format="JPEG", quality=85)
            return base64.b64encode(buffered.getvalue()).decode('utf-8')
    except Exception as e:
        logger.error(f"Error converting file to base64: {str(e)}")
        raise

@api_router.post("/bills/upload", response_model=BillResponse)
async def upload_bill(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload and process bill"""
    if user.get('subscription_plan') == 'free' and user.get('bill_count', 0) >= 20:
        raise HTTPException(status_code=403, detail="Free plan limit reached. Upgrade to Pro for unlimited uploads.")
    
    allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'application/pdf']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Only JPG, PNG, and PDF are allowed.")
    
    try:
        file_content = await file.read()
        file_type = 'pdf' if file.content_type == 'application/pdf' else 'image'
        
        file_id = str(uuid.uuid4())
        file_ext = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
        file_path = UPLOADS_DIR / f"{file_id}.{file_ext}"
        with open(file_path, 'wb') as f:
            f.write(file_content)
        
        image_base64 = convert_to_base64(file_content, file_type)
        extracted_data = await extract_bill_data(image_base64)
        
        bill = {
            "id": file_id,
            "user_id": user['user_id'],
            "file_name": file.filename,
            "file_path": str(file_path),
            "file_type": file_type,
            "upload_date": datetime.now(timezone.utc).isoformat(),
            "ocr_status": "completed",
            "extracted_data": extracted_data.model_dump()
        }
        await db.bills.insert_one(bill)
        
        await db.users.update_one(
            {"user_id": user['user_id']},
            {"$inc": {"bill_count": 1}}
        )
        
        if extracted_data.buyer_name:
            existing_customer = await db.customers.find_one({
                "user_id": user['user_id'],
                "name": extracted_data.buyer_name
            }, {"_id": 0})
            
            if not existing_customer:
                customer = {
                    "id": str(uuid.uuid4()),
                    "user_id": user['user_id'],
                    "name": extracted_data.buyer_name,
                    "gstin": extracted_data.buyer_gstin,
                    "total_purchases": extracted_data.total_amount or 0.0,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.customers.insert_one(customer)
            else:
                await db.customers.update_one(
                    {"id": existing_customer['id']},
                    {"$inc": {"total_purchases": extracted_data.total_amount or 0.0}}
                )
        
        return BillResponse(**bill)
        
    except Exception as e:
        logger.error(f"Error processing bill: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing bill: {str(e)}")

@api_router.get("/bills", response_model=List[BillResponse])
async def get_bills(
    skip: int = 0,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    """Get all bills for user"""
    bills = await db.bills.find(
        {"user_id": user['user_id']},
        {"_id": 0}
    ).sort("upload_date", -1).skip(skip).limit(limit).to_list(limit)
    return [BillResponse(**bill) for bill in bills]

@api_router.get("/bills/{bill_id}", response_model=BillResponse)
async def get_bill(bill_id: str, user: dict = Depends(get_current_user)):
    """Get single bill"""
    bill = await db.bills.find_one({"id": bill_id, "user_id": user['user_id']}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    return BillResponse(**bill)

@api_router.put("/bills/{bill_id}")
async def update_bill(
    bill_id: str,
    extracted_data: BillExtractedData,
    user: dict = Depends(get_current_user)
):
    """Update bill extracted data"""
    result = await db.bills.update_one(
        {"id": bill_id, "user_id": user['user_id']},
        {"$set": {"extracted_data": extracted_data.model_dump()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user['user_id'],
        "action": "update_bill",
        "entity_type": "bill",
        "entity_id": bill_id,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Bill updated successfully"}

@api_router.delete("/bills/{bill_id}")
async def delete_bill(bill_id: str, user: dict = Depends(get_current_user)):
    """Delete bill"""
    bill = await db.bills.find_one({"id": bill_id, "user_id": user['user_id']}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    file_path = Path(bill['file_path'])
    if file_path.exists():
        file_path.unlink()
    
    await db.bills.delete_one({"id": bill_id})
    return {"message": "Bill deleted successfully"}

# ==================== LEDGER ====================

@api_router.get("/ledger")
async def get_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get ledger entries"""
    query = {"user_id": user['user_id']}
    
    bills = await db.bills.find(query, {"_id": 0}).sort("upload_date", -1).to_list(1000)
    
    ledger_entries = []
    for bill in bills:
        data = bill.get('extracted_data', {})
        if data:
            ledger_entries.append({
                "date": bill['upload_date'],
                "customer": data.get('buyer_name', 'N/A'),
                "invoice_number": data.get('invoice_number', 'N/A'),
                "products": len(data.get('products', [])),
                "subtotal": data.get('subtotal', 0),
                "cgst": data.get('cgst', 0),
                "sgst": data.get('sgst', 0),
                "igst": data.get('igst', 0),
                "total_gst": data.get('total_gst', 0),
                "total_amount": data.get('total_amount', 0)
            })
    
    return {"entries": ledger_entries}

@api_router.get("/ledger/export")
async def export_ledger(
    format: str = "xlsx",
    user: dict = Depends(get_current_user)
):
    """Export ledger to Excel or CSV"""
    bills = await db.bills.find({"user_id": user['user_id']}, {"_id": 0}).sort("upload_date", -1).to_list(1000)
    
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger"
        
        headers = ["Date", "Customer", "Invoice No", "Products", "Subtotal", "CGST", "SGST", "IGST", "Total GST", "Total Amount"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for bill in bills:
            data = bill.get('extracted_data', {})
            if data:
                ws.append([
                    bill['upload_date'][:10],
                    data.get('buyer_name', 'N/A'),
                    data.get('invoice_number', 'N/A'),
                    len(data.get('products', [])),
                    data.get('subtotal', 0),
                    data.get('cgst', 0),
                    data.get('sgst', 0),
                    data.get('igst', 0),
                    data.get('total_gst', 0),
                    data.get('total_amount', 0)
                ])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ledger.xlsx"}
        )
    
    return {"message": "CSV export not implemented yet"}

# ==================== CUSTOMERS ====================

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer: CustomerBase, user: dict = Depends(get_current_user)):
    customer_data = customer.model_dump()
    customer_data.update({
        "id": str(uuid.uuid4()),
        "user_id": user['user_id'],
        "total_purchases": 0.0,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    await db.customers.insert_one(customer_data)
    return CustomerResponse(**customer_data)

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(user: dict = Depends(get_current_user)):
    customers = await db.customers.find({"user_id": user['user_id']}, {"_id": 0}).to_list(1000)
    return [CustomerResponse(**c) for c in customers]

@api_router.put("/customers/{customer_id}")
async def update_customer(
    customer_id: str,
    customer: CustomerBase,
    user: dict = Depends(get_current_user)
):
    result = await db.customers.update_one(
        {"id": customer_id, "user_id": user['user_id']},
        {"$set": customer.model_dump(exclude_unset=True)}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer updated successfully"}

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, user: dict = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id, "user_id": user['user_id']}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

# ==================== PRODUCTS ====================

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductBase, user: dict = Depends(get_current_user)):
    product_data = product.model_dump()
    product_data.update({
        "id": str(uuid.uuid4()),
        "user_id": user['user_id'],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    await db.products.insert_one(product_data)
    return ProductResponse(**product_data)

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(user: dict = Depends(get_current_user)):
    products = await db.products.find({"user_id": user['user_id']}, {"_id": 0}).to_list(1000)
    return [ProductResponse(**p) for p in products]

@api_router.put("/products/{product_id}")
async def update_product(
    product_id: str,
    product: ProductBase,
    user: dict = Depends(get_current_user)
):
    result = await db.products.update_one(
        {"id": product_id, "user_id": user['user_id']},
        {"$set": product.model_dump(exclude_unset=True)}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product updated successfully"}

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(get_current_user)):
    result = await db.products.delete_one({"id": product_id, "user_id": user['user_id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}

# ==================== INVOICES ====================

@api_router.post("/invoices", response_model=InvoiceResponse)
async def create_invoice(
    invoice_data: InvoiceCreate,
    user: dict = Depends(get_current_user)
):
    subtotal = sum(item.amount for item in invoice_data.items)
    
    gst_rate = 0.18
    
    if invoice_data.customer_gstin:
        igst = round(subtotal * gst_rate, 2)
        cgst = 0.0
        sgst = 0.0
    else:
        cgst = round(subtotal * (gst_rate / 2), 2)
        sgst = round(subtotal * (gst_rate / 2), 2)
        igst = 0.0
    
    total_gst = cgst + sgst + igst
    total_amount = subtotal + total_gst
    
    invoice_count = await db.invoices.count_documents({"user_id": user['user_id']})
    invoice_number = f"INV-{user['user_id'][:8].upper()}-{invoice_count + 1:04d}"
    
    invoice = {
        "id": str(uuid.uuid4()),
        "user_id": user['user_id'],
        "invoice_number": invoice_number,
        "invoice_date": datetime.now(timezone.utc).isoformat(),
        "customer_id": invoice_data.customer_id,
        "customer_name": invoice_data.customer_name,
        "customer_gstin": invoice_data.customer_gstin,
        "customer_address": invoice_data.customer_address,
        "items": [item.model_dump() for item in invoice_data.items],
        "subtotal": subtotal,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "total_gst": total_gst,
        "total_amount": total_amount,
        "notes": invoice_data.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.invoices.insert_one(invoice)
    return InvoiceResponse(**invoice)

@api_router.get("/invoices", response_model=List[InvoiceResponse])
async def get_invoices(
    skip: int = 0,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    invoices = await db.invoices.find(
        {"user_id": user['user_id']},
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return [InvoiceResponse(**inv) for inv in invoices]

@api_router.get("/invoices/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(invoice_id: str, user: dict = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id, "user_id": user['user_id']}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return InvoiceResponse(**invoice)

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    total_bills = await db.bills.count_documents({"user_id": user['user_id']})
    total_customers = await db.customers.count_documents({"user_id": user['user_id']})
    
    bills = await db.bills.find({"user_id": user['user_id']}, {"_id": 0}).to_list(1000)
    
    total_sales = 0.0
    total_gst = 0.0
    monthly_sales = 0.0
    monthly_gst = 0.0
    
    current_month = datetime.now(timezone.utc).month
    current_year = datetime.now(timezone.utc).year
    
    for bill in bills:
        data = bill.get('extracted_data', {})
        if data:
            amount = data.get('total_amount', 0) or 0
            gst = data.get('total_gst', 0) or 0
            total_sales += amount
            total_gst += gst
            
            bill_date = datetime.fromisoformat(bill['upload_date'])
            if bill_date.month == current_month and bill_date.year == current_year:
                monthly_sales += amount
                monthly_gst += gst
    
    recent_bills_data = await db.bills.find(
        {"user_id": user['user_id']},
        {"_id": 0}
    ).sort("upload_date", -1).limit(5).to_list(5)
    
    recent_bills = [BillResponse(**bill) for bill in recent_bills_data]
    
    return DashboardStats(
        total_bills=total_bills,
        total_customers=total_customers,
        total_sales=round(total_sales, 2),
        total_gst=round(total_gst, 2),
        monthly_sales=round(monthly_sales, 2),
        monthly_gst=round(monthly_gst, 2),
        recent_bills=recent_bills
    )

# ==================== SUBSCRIPTION ====================

@api_router.post("/subscription/create-order")
async def create_subscription_order(
    order: SubscriptionOrder,
    user: dict = Depends(get_current_user)
):
    """Create Razorpay order for subscription"""
    if not razorpay_client:
        raise HTTPException(status_code=500, detail="Payment gateway not configured")
    
    pricing = {
        "pro": {"monthly": 49900, "yearly": 499900},
        "business": {"monthly": 99900, "yearly": 999900}
    }
    
    amount = pricing.get(order.plan, {}).get(order.billing_cycle, 0)
    if amount == 0:
        raise HTTPException(status_code=400, detail="Invalid plan or billing cycle")
    
    try:
        razorpay_order = razorpay_client.order.create({
            "amount": amount,
            "currency": "INR",
            "payment_capture": 1
        })
        
        transaction = {
            "id": str(uuid.uuid4()),
            "user_id": user['user_id'],
            "order_id": razorpay_order['id'],
            "plan": order.plan,
            "billing_cycle": order.billing_cycle,
            "amount": amount,
            "status": "created",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.transactions.insert_one(transaction)
        
        return {
            "order_id": razorpay_order['id'],
            "amount": amount,
            "currency": "INR",
            "key_id": razorpay_key_id
        }
    except Exception as e:
        logger.error(f"Error creating Razorpay order: {str(e)}")
        raise HTTPException(status_code=500, detail="Error creating payment order")

@api_router.post("/subscription/verify")
async def verify_subscription(
    payment_id: str = Form(...),
    order_id: str = Form(...),
    signature: str = Form(...),
    user: dict = Depends(get_current_user)
):
    """Verify Razorpay payment"""
    if not razorpay_client:
        raise HTTPException(status_code=500, detail="Payment gateway not configured")
    
    try:
        razorpay_client.utility.verify_payment_signature({
            'razorpay_order_id': order_id,
            'razorpay_payment_id': payment_id,
            'razorpay_signature': signature
        })
        
        transaction = await db.transactions.find_one({"order_id": order_id}, {"_id": 0})
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        await db.transactions.update_one(
            {"order_id": order_id},
            {"$set": {"status": "completed", "payment_id": payment_id}}
        )
        
        await db.users.update_one(
            {"user_id": user['user_id']},
            {"$set": {"subscription_plan": transaction['plan']}}
        )
        
        return {"message": "Subscription activated successfully", "plan": transaction['plan']}
        
    except razorpay.errors.SignatureVerificationError:
        raise HTTPException(status_code=400, detail="Invalid payment signature")
    except Exception as e:
        logger.error(f"Error verifying payment: {str(e)}")
        raise HTTPException(status_code=500, detail="Error verifying payment")

# Include router
app.include_router(api_router)

# Serve uploaded files
@app.get(\"/uploads/{filename}\")
async def get_upload(filename: str):
    \"\"\"Serve uploaded files\"\"\"
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail=\"File not found\")
    return FileResponse(file_path)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()